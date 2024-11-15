import json
from typing import List, Dict, Any
from openpyxl import load_workbook
from app.utils.pii_scan.structured_ner_main import MLBasedNERScannerForStructuredData  # Adjust import according to your actual module



def read_all_sheets(file_path: str) -> Dict[str, List[Dict[str, str]]]:
    """
    Read all sheets in an Excel (.xlsx) file and return their content as a dictionary of sheet names to lists of dictionaries.
    
    Args:
        file_path (str): Path to the Excel file.
    
    Returns:
        Dict[str, List[Dict[str, str]]]: Dictionary where keys are sheet names and values are lists of row dictionaries.
    """
    workbook = load_workbook(filename=file_path, data_only=True)
    all_sheets_data = {}
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Get headers from the first row
        headers = [cell.value for cell in sheet[1]]
        
        # Extract rows into a list of dictionaries
        sheet_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            sheet_data.append(row_dict)
        
        all_sheets_data[sheet_name] = sheet_data

    return all_sheets_data

def extract_column_data(data: List[Dict[str, str]], column_name: str) -> List[str]:

    return [row.get(column_name, '') for row in data]

def clean_data(data: List[str]) -> List[str]:

    return [text.strip() for text in data if text and isinstance(text, str) and text.strip()]


def process_column_data(column_data: List[str], column_name: str, scanner: MLBasedNERScannerForStructuredData, chunk_size: int = 100) -> Dict[str, Any]:
    """
    Process column data using NER Scanner.
    
    Args:
        column_data (List[str]): List of column data.
        column_name (str): Name of the column.
        scanner (MLBasedNERScannerForStructuredData): NER Scanner instance.
        chunk_size (int): Size of data chunks for processing.
    
    Returns:
        Dict[str, Any]: NER scan result for the column.
    """
    return scanner.scan(column_data, column_name=column_name, chunk_size=chunk_size)

